"""
MedPark 미수채권 관리 — Flask 백엔드.

실행:
    개발  python app.py
    운영  gunicorn -w 2 -b 0.0.0.0:$PORT app:app
"""
import json
import os
import base64
import io
import calendar
from copy import copy
from datetime import date, datetime
from functools import wraps

from flask import Flask, jsonify, request, session, render_template, send_file
from werkzeug.security import check_password_hash, generate_password_hash

import db
from db import connect, PERMISSIONS, ALL_PERMS, ROLE_TEMPLATES

app = Flask(__name__, static_folder="static", template_folder="templates")
app.secret_key = db.secret_key()
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    MAX_CONTENT_LENGTH=24 * 1024 * 1024,
)

app.json.ensure_ascii = False

db.init_db()

METHODS = ["계좌수금", "카드수금", "어음수금", "현금수금", "기타수금"]
UNITS = ["덴탈", "메디컬", "에스테틱"]
STATUSES = ["정상", "연체", "부실"]


# ─────────────────────────────── 인증 ───────────────────────────────

def current_user():
    username = session.get("username")
    if not username:
        return None
    with connect() as conn:
        row = conn.execute(
            "SELECT username, name, title, role, biz_unit, permissions, active"
            " FROM users WHERE username = %s", (username,)).fetchone()
    if not row or not row["active"]:
        return None
    user = dict(row)
    user["permissions"] = json.loads(user["permissions"])
    return user


def login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        user = current_user()
        if not user:
            return jsonify(error="로그인이 필요합니다."), 401
        request.user = user
        return fn(*args, **kwargs)
    return wrapper


def requires(perm):
    def decorator(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            user = current_user()
            if not user:
                return jsonify(error="로그인이 필요합니다."), 401
            if perm not in user["permissions"]:
                label = dict(PERMISSIONS).get(perm, perm)
                return jsonify(error="'%s' 권한이 없습니다. 관리자에게 요청하세요." % label), 403
            request.user = user
            return fn(*args, **kwargs)
        return wrapper
    return decorator


def log(conn, actor, action, detail=""):
    conn.execute("INSERT INTO audit (actor, action, detail) VALUES (%s,%s,%s)",
                 (actor, action, detail))


def body():
    return request.get_json(silent=True) or {}


def as_int(value, default=0):
    try:
        return int(round(float(str(value).replace(",", "").strip() or 0)))
    except (TypeError, ValueError):
        return default


def add_months(month, offset):
    year, mon = (int(x) for x in month.split("-"))
    total = year * 12 + mon - 1 + int(offset)
    return "%04d-%02d" % (total // 12, total % 12 + 1)


def month_offset(base_month, target_month):
    by, bm = (int(x) for x in base_month.split("-"))
    ty, tm = (int(x) for x in target_month.split("-"))
    return (ty - by) * 12 + tm - bm


# ─────────────────────────────── 화면 ───────────────────────────────

BUILD = str(int(os.path.getmtime(os.path.join(os.path.dirname(__file__), "static", "app.js"))))


@app.get("/")
def index():
    # build 값을 쿼리스트링에 붙여 배포 후 브라우저 캐시가 남지 않게 한다.
    return render_template("index.html", build=BUILD)


@app.get("/health")
def health():
    return jsonify(status="ok", db=db.backend_name(), journal=db.journal_mode(),
                   time=datetime.now().isoformat(timespec="seconds"))


# ─────────────────────────────── 세션 API ───────────────────────────

@app.post("/api/login")
def login():
    data = body()
    username = (data.get("username") or "").strip()
    password = data.get("password") or ""
    with connect() as conn:
        row = conn.execute("SELECT * FROM users WHERE username = %s", (username,)).fetchone()
        if not row or not check_password_hash(row["password"], password):
            return jsonify(error="아이디 또는 비밀번호가 일치하지 않습니다."), 401
        if not row["active"]:
            return jsonify(error="비활성 계정입니다. 관리자에게 문의하세요."), 403
        log(conn, username, "login")
    session["username"] = username
    session.permanent = True
    return jsonify(user=current_user())


@app.post("/api/logout")
def logout():
    session.clear()
    return jsonify(ok=True)


@app.get("/api/me")
def me():
    user = current_user()
    return jsonify(user=user) if user else (jsonify(user=None), 200)


@app.post("/api/password")
@login_required
def change_password():
    data = body()
    new = data.get("password") or ""
    if len(new) < 8:
        return jsonify(error="비밀번호는 8자 이상이어야 합니다."), 400
    with connect() as conn:
        row = conn.execute("SELECT password FROM users WHERE username = %s",
                           (request.user["username"],)).fetchone()
        if not check_password_hash(row["password"], data.get("current") or ""):
            return jsonify(error="현재 비밀번호가 일치하지 않습니다."), 400
        conn.execute("UPDATE users SET password = %s WHERE username = %s",
                     (generate_password_hash(new), request.user["username"]))
        log(conn, request.user["username"], "password_change")
    return jsonify(ok=True)


# ─────────────────────────────── 부트스트랩 ─────────────────────────

@app.get("/api/bootstrap")
@login_required
def bootstrap():
    user = request.user
    with connect() as conn:
        customers = [r for r in conn.execute(
            "SELECT * FROM customers ORDER BY balance DESC")]
        collections = [r for r in conn.execute(
            "SELECT * FROM collections ORDER BY id DESC LIMIT 800")]
        targets = [r for r in conn.execute(
            "SELECT * FROM targets ORDER BY target_date ASC LIMIT 800")]
        uploads = [r for r in conn.execute(
            "SELECT * FROM uploads ORDER BY id DESC LIMIT 200")]
        locks = [r for r in conn.execute("SELECT * FROM month_locks")]
        finalized = sorted([r["month"] for r in locks if r.get("locked")])
        latest_snapshot = next((u["month"] for u in uploads
                                if u.get("upload_type") == "snapshot"), date.today().strftime("%Y-%m"))
        cash_plan_months = [add_months(latest_snapshot, i) for i in range(3)]
        reflection_label = "마감 데이터 없음"
        if finalized:
            closed_month = finalized[-1]
            reflection_label = "%d월 최종마감 반영" % int(closed_month[5:7])
            next_month = add_months(closed_month, 1)
            shipment = next((u for u in uploads
                             if u.get("upload_type") == "shipment" and u["month"] == next_month), None)
            if shipment:
                reflected_date = shipment.get("shipment_date") or str(shipment["uploaded_at"])[:10]
                uploaded_day = int(reflected_date[8:10])
                reflection_label += " + %d월 %d일 출고데이터 반영" % (
                    int(next_month[5:7]), uploaded_day)
        users = []
        if "user_manage" in user["permissions"]:
            users = [{k: v for k, v in r.items() if k != "password"}
                     for r in conn.execute("SELECT * FROM users ORDER BY username")]
            for u in users:
                u["permissions"] = json.loads(u["permissions"])
    return jsonify(
        user=user, customers=customers, collections=collections, targets=targets,
        uploads=uploads, locks=locks, users=users,
        meta=dict(permissions=[{"key": k, "label": l} for k, l in PERMISSIONS],
                  roles={k: v for k, v in ROLE_TEMPLATES.items()},
                  methods=METHODS, units=UNITS, statuses=STATUSES,
                  today=date.today().isoformat(), reflection_label=reflection_label,
                  cash_plan_months=cash_plan_months),
    )


# ─────────────────────────────── 거래처 ─────────────────────────────

@app.patch("/api/customers/<code>")
@login_required
def update_customer(code):
    data = body()
    fields, values = [], []
    if "period" in data:
        if "customer_info_edit" not in request.user["permissions"]:
            return jsonify(error="'거래처 정보수정' 권한이 없습니다."), 403
        period = as_int(data.get("period"), -1)
        if period < 0:
            return jsonify(error="회수기간은 0 이상의 개월 수로 입력하세요."), 400
        fields.append("period = %s")
        values.append(period)
    for key in ("note", "owner", "status", "collection_target_date"):
        if key in data:
            if "note_edit" not in request.user["permissions"]:
                return jsonify(error="'비고 편집' 권한이 없습니다."), 403
            fields.append(key + " = %s")
            values.append(str(data[key]))
    if not fields:
        return jsonify(error="변경할 항목이 없습니다."), 400
    values.append(code)
    with connect() as conn:
        cur = conn.execute(
            "UPDATE customers SET " + ", ".join(fields)
            + ", updated_at = " + db.NOW_SQL + " WHERE code = %s", values)
        if cur.rowcount == 0:
            return jsonify(error="거래처를 찾을 수 없습니다."), 404
        log(conn, request.user["username"], "customer_update", code)
        row = conn.execute("SELECT * FROM customers WHERE code = %s", (code,)).fetchone()
    return jsonify(customer=row)


# ─────────────────────────────── 수금 ───────────────────────────────

@app.post("/api/collections")
@requires("collection_register")
def register_collection():
    data = body()
    code = (data.get("customer_code") or "").strip()
    amount = as_int(data.get("amount"))
    method = data.get("method") or ""
    paid_at = data.get("paid_at") or date.today().isoformat()
    if not code:
        return jsonify(error="거래처를 선택하세요."), 400
    if amount <= 0:
        return jsonify(error="수금액은 0보다 커야 합니다."), 400
    if method not in METHODS:
        return jsonify(error="수금방법을 선택하세요."), 400
    with connect() as conn:
        cust = conn.execute("SELECT name, balance FROM customers WHERE code = %s",
                            (code,)).fetchone()
        if not cust:
            return jsonify(error="등록되지 않은 거래처입니다."), 404
        new_id = conn.execute(
            "INSERT INTO collections (customer_code, customer_name, amount, method,"
            " paid_at, state, registered_by, note) VALUES (%s,%s,%s,%s,%s,'pending',%s,%s)"
            " RETURNING id",
            (code, cust["name"], amount, method, paid_at,
             request.user["username"], data.get("note") or "")).fetchone()["id"]
        log(conn, request.user["username"], "collection_register", "%s / %d" % (code, amount))
        row = conn.execute("SELECT * FROM collections WHERE id = %s", (new_id,)).fetchone()
    return jsonify(collection=row), 201


@app.post("/api/collections/<int:cid>/approve")
@requires("collection_approve")
def approve_collection(cid):
    """
    승인은 잔액을 깎는 동작이므로 한 건이 두 번 반영되면 안 된다.
    상태를 조회한 뒤 갱신하면 그 사이에 다른 요청이 끼어들 수 있으므로,
    'pending 인 경우에만' 이라는 조건을 UPDATE 문 안에 넣어 한 문장으로 처리한다.
    조건에 걸려 0행이 바뀌면 다른 요청이 이미 가져간 것이다.
    """
    with connect() as conn:
        before = conn.execute(
            "SELECT bad_balance, overdue_balance, normal_balance FROM customers WHERE code=("
            "SELECT customer_code FROM collections WHERE id=%s)", (cid,)).fetchone()
        row = conn.execute(
            "UPDATE collections SET state='approved', approved_by=%s,"
            " approved_at=" + db.NOW_SQL +
            " WHERE id=%s AND state='pending' RETURNING *",
            (request.user["username"], cid)).fetchone()
        if row is None:
            exists = conn.execute("SELECT state FROM collections WHERE id = %s",
                                  (cid,)).fetchone()
            if not exists:
                return jsonify(error="수금 건을 찾을 수 없습니다."), 404
            return jsonify(error="이미 처리된 건입니다."), 409

        # 잔액 차감도 읽고 쓰지 않고 한 문장으로 끝낸다.
        amount = row["amount"]
        normal_paid = 0
        if before:
            normal_paid = min(
                max(amount - before["bad_balance"] - before["overdue_balance"], 0),
                before["normal_balance"])
        customer = conn.execute(
            "UPDATE customers SET"
            " balance = CASE WHEN balance - %s < 0 THEN 0 ELSE balance - %s END,"
            " bad_balance = CASE WHEN bad_balance >= %s THEN bad_balance - %s ELSE 0 END,"
            " overdue_balance = CASE"
            "   WHEN %s <= bad_balance THEN overdue_balance"
            "   WHEN %s - bad_balance <= overdue_balance THEN overdue_balance - (%s - bad_balance)"
            "   ELSE 0 END,"
            " normal_balance = CASE"
            "   WHEN %s <= bad_balance + overdue_balance THEN normal_balance"
            "   WHEN %s - bad_balance - overdue_balance <= normal_balance"
            "     THEN normal_balance - (%s - bad_balance - overdue_balance)"
            "   ELSE 0 END,"
            " status = CASE WHEN balance - %s <= 0 THEN '정상' ELSE status END,"
            " overdue_days = CASE WHEN balance - %s <= 0 THEN 0 ELSE overdue_days END,"
            " last_paid_at = %s, updated_at = " + db.NOW_SQL +
            " WHERE code = %s RETURNING *",
            (amount, amount, amount, amount,
             amount, amount, amount,
             amount, amount, amount,
             amount, amount, row["paid_at"], row["customer_code"])).fetchone()
        # 정상채권까지 차감된 경우 수금대상월이 오래된 출고분부터 잔액을 줄인다.
        remaining = normal_paid
        if remaining > 0:
            shipment_rows = [s for s in conn.execute(
                "SELECT month,code,balance FROM monthly_shipments"
                " WHERE code=%s AND balance>0 ORDER BY target_month,month",
                (row["customer_code"],))]
            for shipment in shipment_rows:
                paid = min(remaining, shipment["balance"])
                conn.execute(
                    "UPDATE monthly_shipments SET balance=balance-%s WHERE month=%s AND code=%s",
                    (paid, shipment["month"], shipment["code"]))
                remaining -= paid
                if remaining <= 0:
                    break
        log(conn, request.user["username"], "collection_approve", str(cid))
    return jsonify(collection=row, customer=customer)


@app.post("/api/collections/<int:cid>/reject")
@requires("collection_approve")
def reject_collection(cid):
    reason = (body().get("reason") or "").strip()
    with connect() as conn:
        collection = conn.execute(
            "UPDATE collections SET state='rejected', approved_by=%s,"
            " approved_at=" + db.NOW_SQL + ", reject_reason=%s"
            " WHERE id=%s AND state='pending' RETURNING *",
            (request.user["username"], reason, cid)).fetchone()
        if collection is None:
            exists = conn.execute("SELECT state FROM collections WHERE id = %s",
                                  (cid,)).fetchone()
            if not exists:
                return jsonify(error="수금 건을 찾을 수 없습니다."), 404
            return jsonify(error="이미 처리된 건입니다."), 409
        log(conn, request.user["username"], "collection_reject", str(cid))
    return jsonify(collection=collection)


# ─────────────────────────────── 수금목표 ───────────────────────────

@app.post("/api/targets")
@requires("target_manage")
def create_target():
    data = body()
    code = (data.get("customer_code") or "").strip()
    target_date = data.get("target_date") or ""
    if not code or not target_date:
        return jsonify(error="거래처와 목표일을 입력하세요."), 400
    with connect() as conn:
        cust = conn.execute("SELECT name FROM customers WHERE code=%s", (code,)).fetchone()
        if not cust:
            return jsonify(error="등록되지 않은 거래처입니다."), 404
        new_id = conn.execute(
            "INSERT INTO targets (customer_code, customer_name, amount, target_date,"
            " method, assignee, note, created_by) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)"
            " RETURNING id",
            (code, cust["name"], as_int(data.get("amount")), target_date,
             data.get("method") or "", data.get("assignee") or "",
             data.get("note") or "", request.user["username"])).fetchone()["id"]
        row = conn.execute("SELECT * FROM targets WHERE id=%s", (new_id,)).fetchone()
    return jsonify(target=row), 201


@app.patch("/api/targets/<int:tid>")
@requires("target_manage")
def update_target(tid):
    data = body()
    allowed = ("amount", "target_date", "done_date", "method", "assignee", "note", "state")
    fields, values = [], []
    for key in allowed:
        if key in data:
            fields.append(key + " = %s")
            values.append(as_int(data[key]) if key == "amount" else str(data[key]))
    if not fields:
        return jsonify(error="변경할 항목이 없습니다."), 400
    if data.get("done_date") and "state" not in data:
        fields.append("state = %s")
        values.append("done")
    values.append(tid)
    with connect() as conn:
        cur = conn.execute("UPDATE targets SET " + ", ".join(fields) + " WHERE id = %s", values)
        if cur.rowcount == 0:
            return jsonify(error="목표를 찾을 수 없습니다."), 404
        row = conn.execute("SELECT * FROM targets WHERE id=%s", (tid,)).fetchone()
    return jsonify(target=row)


@app.delete("/api/targets/<int:tid>")
@requires("target_manage")
def delete_target(tid):
    with connect() as conn:
        conn.execute("DELETE FROM targets WHERE id = %s", (tid,))
    return jsonify(ok=True)


# ─────────────────────────────── 업로드 ─────────────────────────────

@app.post("/api/uploads")
@requires("upload_data")
def upload_rows():
    """클라이언트(SheetJS)가 파싱한 행을 받아 해당 월 파티션만 교체한다."""
    data = body()
    month = (data.get("month") or "").strip()
    rows = data.get("rows") or []
    upload_type = data.get("mode") or "snapshot"
    shipment_date = (data.get("shipment_date") or "").strip()
    filename = data.get("filename") or "unknown.xlsx"
    if len(month) != 7 or month[4] != "-":
        return jsonify(error="기준월 형식이 올바르지 않습니다. 예: 2026-08"), 400
    if not rows:
        return jsonify(error="읽어들인 행이 없습니다. 시트와 머리글을 확인하세요."), 400

    with connect() as conn:
        lock = conn.execute("SELECT locked FROM month_locks WHERE month = %s", (month,)).fetchone()
        if lock and lock["locked"]:
            return jsonify(error="%s 은 마감 잠금 상태입니다. 잠금을 해제한 뒤 업로드하세요." % month), 423

        if upload_type == "shipment":
            try:
                parsed_shipment_date = datetime.strptime(shipment_date, "%Y-%m-%d").date()
            except ValueError:
                return jsonify(error="출고기준일을 입력하세요."), 400
            if parsed_shipment_date.strftime("%Y-%m") != month:
                return jsonify(error="출고기준일은 선택한 기준월 안의 날짜여야 합니다."), 400
            shipments = {}
            for r in rows:
                code = str(r.get("code") or "").strip()
                if not code:
                    continue
                if code.isdigit():
                    code = code.zfill(5)
                existing = conn.execute("SELECT period FROM customers WHERE code=%s", (code,)).fetchone()
                period_raw = r.get("collection_period")
                period = (existing["period"] if period_raw in (None, "") and existing
                          else as_int(period_raw, -1))
                amount = as_int(r.get("shipment_amount"))
                if amount < 0:
                    return jsonify(error="%s 거래처의 출고금액은 0 이상이어야 합니다." % code), 400
                target_month = add_months(month, period) if period >= 0 else ""
                bucket = "current" if period == 0 else ("next" if period == 1 else "later")
                shipments[code] = dict(
                    code=code, name=str(r.get("name") or "").strip() or code,
                    biz_unit=str(r.get("biz_unit") or "").strip() or "덴탈",
                    owner=str(r.get("owner") or "").strip(), period=period,
                    target_month=target_month, bucket=bucket, amount=amount,
                    note=str(r.get("note") or "").strip())

            previous_rows = [x for x in conn.execute(
                "SELECT * FROM monthly_shipments WHERE month=%s", (month,))]
            previous = len(previous_rows)
            bucket_columns = {
                "current": "normal_current_balance",
                "next": "normal_next_balance",
                "later": "normal_later_balance",
            }
            # 같은 월 재업로드는 그 월 출고분의 남은 잔액만 먼저 제거한다.
            for old in previous_rows:
                amount = old["balance"]
                column = bucket_columns.get(old["bucket"], "normal_later_balance")
                conn.execute(
                    "UPDATE customers SET balance=CASE WHEN balance-%s<0 THEN 0 ELSE balance-%s END,"
                    " normal_balance=CASE WHEN normal_balance-%s<0 THEN 0 ELSE normal_balance-%s END,"
                    " " + column + "=CASE WHEN " + column + "-%s<0 THEN 0 ELSE " + column + "-%s END"
                    " WHERE code=%s",
                    (amount, amount, amount, amount, amount, amount, old["code"]))
            conn.execute("DELETE FROM monthly_shipments WHERE month=%s", (month,))

            for item in shipments.values():
                current = conn.execute("SELECT 1 FROM customers WHERE code=%s", (item["code"],)).fetchone()
                column = bucket_columns[item["bucket"]]
                if current:
                    conn.execute(
                        "UPDATE customers SET name=%s, biz_unit=%s,"
                        " owner=CASE WHEN %s='' THEN owner ELSE %s END, period=%s,"
                        " balance=balance+%s, normal_balance=normal_balance+%s,"
                        " " + column + "=" + column + "+%s, source_month=%s,"
                        " updated_at=" + db.NOW_SQL + " WHERE code=%s",
                        (item["name"], item["biz_unit"], item["owner"], item["owner"],
                         item["period"], item["amount"], item["amount"], item["amount"],
                         month, item["code"]))
                else:
                    values = dict(current=0, next=0, later=0)
                    values[item["bucket"]] = item["amount"]
                    conn.execute(
                        "INSERT INTO customers (code,name,biz_unit,status,owner,balance,normal_balance,"
                        " normal_current_balance,normal_next_balance,normal_later_balance,period,source_month,note)"
                        " VALUES (%s,%s,%s,'정상',%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                        (item["code"], item["name"], item["biz_unit"], item["owner"],
                         item["amount"], item["amount"], values["current"], values["next"],
                         values["later"], item["period"], month, item["note"]))
                conn.execute(
                    "INSERT INTO monthly_shipments (month,code,name,biz_unit,owner,collection_period,"
                    " target_month,bucket,amount,balance,note) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                    (month, item["code"], item["name"], item["biz_unit"], item["owner"],
                     item["period"], item["target_month"], item["bucket"], item["amount"],
                     item["amount"], item["note"]))
            conn.execute(
                "INSERT INTO uploads (month,filename,row_count,uploaded_by,replaced,upload_type,shipment_date)"
                " VALUES (%s,%s,%s,%s,%s,'shipment',%s)",
                (month, filename, len(shipments), request.user["username"], previous, shipment_date))
            log(conn, request.user["username"], "shipment_upload",
                "%s / %d행 (기존 %d행 교체)" % (month, len(shipments), previous))
            customers = [x for x in conn.execute("SELECT * FROM customers ORDER BY balance DESC")]
            uploads = [x for x in conn.execute("SELECT * FROM uploads ORDER BY id DESC LIMIT 200")]
            return jsonify(inserted=len(shipments), replaced=previous,
                           customers=customers, uploads=uploads)

        prev = conn.execute(
            "SELECT COUNT(*) AS c FROM customers WHERE source_month = %s", (month,)).fetchone()["c"]
        conn.execute("DELETE FROM customers WHERE source_month = %s", (month,))

        payload = {}   # 같은 코드가 여러 번 오면 마지막 값만 남긴다
        for r in rows:
            code = str(r.get("code") or "").strip()
            if not code or code.startswith("#REF") or code.startswith("#N/A"):
                continue  # 엑셀 수식 오류 셀은 건너뛴다
            if code.isdigit():
                code = code.zfill(5)
            unit = str(r.get("biz_unit") or "").strip() or "덴탈"
            overdue = as_int(r.get("overdue_days"))
            normal_balance = as_int(r.get("normal_balance"))
            overdue_balance = as_int(r.get("overdue_balance"))
            bad_balance = as_int(r.get("bad_balance"))
            split_total = normal_balance + overdue_balance + bad_balance
            balance = as_int(r.get("balance"))
            if split_total and balance != split_total:
                balance = split_total
            normal_later = as_int(r.get("normal_later_balance"))
            normal_next = as_int(r.get("normal_next_balance"))
            normal_current = as_int(r.get("normal_current_balance"))
            if normal_balance and not (normal_later or normal_next or normal_current):
                normal_current = normal_balance
            status = str(r.get("status") or "").strip()
            if status not in STATUSES:
                status = "부실" if bad_balance else (
                    "연체" if overdue_balance or overdue > 0 else "정상")
            period_raw = r.get("collection_period")
            collection_period = -1 if period_raw in (None, "") else as_int(period_raw, -1)
            payload[code] = (
                code, str(r.get("name") or "").strip() or code, unit, status,
                str(r.get("owner") or "").strip(), balance,
                normal_balance,
                normal_later,
                normal_next,
                normal_current,
                as_int(r.get("normal_collected")),
                overdue_balance,
                as_int(r.get("overdue_source_balance") or overdue_balance),
                as_int(r.get("overdue_collected")),
                bad_balance,
                as_int(r.get("advance")), overdue,
                str(r.get("last_paid_at") or "").strip(), collection_period, month,
                str(r.get("note") or "").strip(),
            )
        conn.executemany(
            "INSERT INTO customers (code, name, biz_unit, status, owner, balance,"
            " normal_balance, normal_later_balance, normal_next_balance,"
            " normal_current_balance, normal_collected, overdue_balance,"
            " overdue_source_balance, overdue_collected, bad_balance, advance, overdue_days,"
            " last_paid_at, period, source_month, note)"
            " VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
            " ON CONFLICT (code) DO UPDATE SET"
            " name=excluded.name, biz_unit=excluded.biz_unit, status=excluded.status,"
            " owner=excluded.owner, balance=excluded.balance, advance=excluded.advance,"
            " normal_balance=excluded.normal_balance,"
            " normal_later_balance=excluded.normal_later_balance,"
            " normal_next_balance=excluded.normal_next_balance,"
            " normal_current_balance=excluded.normal_current_balance,"
            " normal_collected=excluded.normal_collected,"
            " overdue_balance=excluded.overdue_balance, bad_balance=excluded.bad_balance,"
            " overdue_source_balance=excluded.overdue_source_balance,"
            " overdue_collected=excluded.overdue_collected,"
            " overdue_days=excluded.overdue_days, last_paid_at=excluded.last_paid_at,"
            " period=excluded.period, source_month=excluded.source_month, note=excluded.note",
            list(payload.values()))
        conn.execute(
            "INSERT INTO uploads (month, filename, row_count, uploaded_by, replaced, upload_type)"
            " VALUES (%s,%s,%s,%s,%s,'snapshot')",
            (month, filename, len(payload), request.user["username"], prev))
        log(conn, request.user["username"], "upload",
            "%s / %d행 (기존 %d행 교체)" % (month, len(payload), prev))
        customers = [x for x in conn.execute(
            "SELECT * FROM customers ORDER BY balance DESC")]
        uploads = [x for x in conn.execute(
            "SELECT * FROM uploads ORDER BY id DESC LIMIT 200")]
    return jsonify(inserted=len(payload), replaced=prev,
                   customers=customers, uploads=uploads)


# ───────────────────────── 자금수지 수금계획 ──────────────────────

@app.post("/api/cash-plan/export")
@requires("data_export")
def export_cash_plan():
    """첨부된 ㈜메드파크 자금수지 서식을 유지해 월별 수금계획을 생성한다."""
    from openpyxl import load_workbook

    data = body()
    month = (data.get("month") or "").strip()
    include_overdue = bool(data.get("include_overdue"))
    if len(month) != 7 or month[4] != "-":
        return jsonify(error="다운로드 기준월을 선택하세요."), 400

    with connect() as conn:
        customers = [r for r in conn.execute(
            "SELECT * FROM customers ORDER BY biz_unit,name,code")]
        snapshot = conn.execute(
            "SELECT month FROM uploads WHERE upload_type='snapshot' ORDER BY id DESC LIMIT 1"
        ).fetchone()
    base_month = snapshot["month"] if snapshot else month
    offset = month_offset(base_month, month)
    if offset not in (0, 1, 2):
        return jsonify(error="현재 수금계획은 %s부터 3개월 범위에서 다운로드할 수 있습니다." % base_month), 400
    normal_field = ("normal_current_balance" if offset == 0 else
                    "normal_next_balance" if offset == 1 else
                    "normal_later_balance" if offset >= 2 else None)

    rows = []
    for customer in customers:
        if normal_field and customer[normal_field] > 0:
            rows.append((customer, "정상", customer[normal_field]))
        if include_overdue and customer["overdue_balance"] > 0:
            rows.append((customer, "연체", customer["overdue_balance"]))
        if include_overdue and customer["bad_balance"] > 0:
            rows.append((customer, "부실", customer["bad_balance"]))

    template_path = os.path.join(app.static_folder, "cash_plan_template.b64")
    with open(template_path, "rb") as fh:
        template_bytes = base64.b64decode(fh.read())
    wb = load_workbook(io.BytesIO(template_bytes))
    ws = wb["서식"] if "서식" in wb.sheetnames else wb.active

    required_rows = max(len(rows), 16)
    if required_rows > 16:
        ws.insert_rows(23, required_rows - 16)
        for target_row in range(23, 7 + required_rows):
            ws.row_dimensions[target_row].height = ws.row_dimensions[22].height
            for col in range(1, 44):
                target = ws.cell(target_row, col)
                target._style = copy(ws.cell(22, col)._style)

    for row_no in range(7, 7 + required_rows):
        for col in range(1, 44):
            ws.cell(row_no, col).value = None

    year, mon = (int(x) for x in month.split("-"))
    month_end = date(year, mon, calendar.monthrange(year, mon)[1])
    dept = {"덴탈": "국내_덴탈", "메디컬": "국내_메디컬", "에스테틱": "국내_에스테틱"}
    execution = {k: v + " 수금" for k, v in dept.items()}
    detail = {
        "정상": "제품구매대금(%02d월 정상채권)" % mon,
        "연체": "제품구매대금(미수채권)",
        "부실": "제품구매대금(부실채권)",
    }
    for row_no, (customer, category, amount) in enumerate(rows, start=7):
        target_text = (customer.get("collection_target_date") or "").strip()
        try:
            plan_date = datetime.strptime(target_text, "%Y-%m-%d").date() if target_text else month_end
        except ValueError:
            plan_date = month_end
        unit = customer["biz_unit"]
        values = {
            1: "예정", 5: "사업부", 6: dept.get(unit, unit), 7: "수금",
            8: execution.get(unit, unit + " 수금"), 10: plan_date, 12: plan_date,
            13: customer["name"], 14: detail[category], 17: amount,
        }
        for col, value in values.items():
            ws.cell(row_no, col).value = value
        ws.cell(row_no, 10).number_format = "yyyy-mm-dd"
        ws.cell(row_no, 12).number_format = "yyyy-mm-dd"
        ws.cell(row_no, 17).number_format = "#,##0_);[Red](#,##0)"

    ws["F4"] = date.today()
    ws["F4"].number_format = "yyyy-mm-dd"
    ws["F3"] = date.fromordinal(date.today().toordinal() - 1)
    ws["F3"].number_format = "yyyy-mm-dd"
    ws["I1"] = "★ %d/%d 수금계획" % (mon, month_end.day)
    ws.auto_filter.ref = "A6:AQ%d" % max(7, 6 + len(rows))
    ws.print_area = "A1:AQ%d" % max(22, 6 + len(rows))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    suffix = "_미수부실포함" if include_overdue else ""
    filename = "MedPark_%02d월_수금계획%s.xlsx" % (mon, suffix)
    return send_file(output,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=filename)


@app.post("/api/locks/<month>")
@requires("month_lock")
def toggle_lock(month):
    locked = 1 if body().get("locked") else 0
    with connect() as conn:
        conn.execute(
            "INSERT INTO month_locks (month, locked, locked_by, locked_at)"
            " VALUES (%s,%s,%s," + db.NOW_SQL + ")"
            " ON CONFLICT(month) DO UPDATE SET locked=excluded.locked,"
            " locked_by=excluded.locked_by, locked_at=excluded.locked_at",
            (month, locked, request.user["username"]))
        log(conn, request.user["username"], "month_lock", "%s → %s" % (month, locked))
        locks = [r for r in conn.execute("SELECT * FROM month_locks")]
    return jsonify(locks=locks)


# ─────────────────────────────── 계정·권한 ──────────────────────────

@app.post("/api/users")
@requires("user_manage")
def create_user():
    data = body()
    username = (data.get("username") or "").strip()
    password = data.get("password") or ""
    if not username:
        return jsonify(error="아이디를 입력하세요."), 400
    if len(password) < 8:
        return jsonify(error="초기 비밀번호는 8자 이상이어야 합니다."), 400
    role = data.get("role") if data.get("role") in ROLE_TEMPLATES else "sales"
    perms = data.get("permissions")
    if not isinstance(perms, list):
        perms = ROLE_TEMPLATES[role]["perms"]
    perms = [p for p in perms if p in ALL_PERMS]
    with connect() as conn:
        if conn.execute("SELECT 1 FROM users WHERE username=%s", (username,)).fetchone():
            return jsonify(error="이미 존재하는 아이디입니다."), 409
        conn.execute(
            "INSERT INTO users (username, name, title, role, biz_unit, password, permissions)"
            " VALUES (%s,%s,%s,%s,%s,%s,%s)",
            (username, data.get("name") or username, data.get("title") or "", role,
             data.get("biz_unit") or "",
             generate_password_hash(password),
             json.dumps(perms, ensure_ascii=False)))
        log(conn, request.user["username"], "user_create", username)
    return jsonify(ok=True), 201


@app.patch("/api/users/<username>")
@requires("user_manage")
def update_user(username):
    data = body()
    fields, values = [], []
    for key in ("name", "title", "biz_unit"):
        if key in data:
            fields.append(key + " = %s")
            values.append(str(data[key]))
    if data.get("role") in ROLE_TEMPLATES:
        fields.append("role = %s")
        values.append(data["role"])
    if isinstance(data.get("permissions"), list):
        fields.append("permissions = %s")
        values.append(json.dumps([p for p in data["permissions"] if p in ALL_PERMS],
                                 ensure_ascii=False))
    if "active" in data:
        fields.append("active = %s")
        values.append(1 if data["active"] else 0)
    if data.get("password"):
        fields.append("password = %s")
        values.append(generate_password_hash(data["password"]))
    if not fields:
        return jsonify(error="변경할 항목이 없습니다."), 400
    values.append(username)
    with connect() as conn:
        if username == "Medpark0" and data.get("active") is False:
            return jsonify(error="시스템관리자 계정은 비활성화할 수 없습니다."), 400
        cur = conn.execute("UPDATE users SET " + ", ".join(fields) + " WHERE username = %s", values)
        if cur.rowcount == 0:
            return jsonify(error="계정을 찾을 수 없습니다."), 404
        log(conn, request.user["username"], "user_update", username)
        users = [{k: v for k, v in r.items() if k != "password"}
                 for r in conn.execute("SELECT * FROM users ORDER BY username")]
        for u in users:
            u["permissions"] = json.loads(u["permissions"])
    return jsonify(users=users)


# ─────────────────── window.storage 호환 계층 ───────────────────────
# 기존 아티팩트 코드가 window.storage 를 그대로 호출해도 동작하도록 유지한다.

def _scope(shared):
    return "shared" if shared else "user:%s" % session.get("username", "anon")


@app.get("/api/storage")
@login_required
def storage_get():
    key = request.args.get("key")
    shared = request.args.get("shared") == "true"
    prefix = request.args.get("prefix")
    with connect() as conn:
        if key:
            row = conn.execute("SELECT value FROM kv WHERE scope=%s AND key=%s",
                               (_scope(shared), key)).fetchone()
            if not row:
                return jsonify(error="not found"), 404
            return jsonify(key=key, value=row["value"], shared=shared)
        rows = conn.execute(
            "SELECT key FROM kv WHERE scope=%s AND key LIKE %s ORDER BY key",
            (_scope(shared), (prefix or "") + "%")).fetchall()
    return jsonify(keys=[r["key"] for r in rows], prefix=prefix, shared=shared)


@app.post("/api/storage")
@login_required
def storage_set():
    data = body()
    shared = bool(data.get("shared"))
    with connect() as conn:
        conn.execute(
            "INSERT INTO kv (scope, key, value) VALUES (%s,%s,%s)"
            " ON CONFLICT(scope, key) DO UPDATE SET value = excluded.value",
            (_scope(shared), data.get("key"), json.dumps(data.get("value"))
             if not isinstance(data.get("value"), str) else data["value"]))
    return jsonify(key=data.get("key"), value=data.get("value"), shared=shared)


@app.delete("/api/storage")
@login_required
def storage_delete():
    key = request.args.get("key")
    shared = request.args.get("shared") == "true"
    with connect() as conn:
        conn.execute("DELETE FROM kv WHERE scope=%s AND key=%s", (_scope(shared), key))
    return jsonify(key=key, deleted=True, shared=shared)


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8000))
    app.run(host="0.0.0.0", port=port, debug=False)
